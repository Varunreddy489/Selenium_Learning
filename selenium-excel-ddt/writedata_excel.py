import openpyxl

path = 'Book1.xlsx'

# file -> workbook -> sheet -> rows -> cells

workbook = openpyxl.load_workbook(path)
sheet = workbook.active  # get active sheet

# add same data
# for r in range(1, 6):
#     for c in range(1,4):
#         sheet.cell(r, c).value="Welcome"


# add multiple data

sheet.cell(1,1).value=1
sheet.cell(1,2).value=2
sheet.cell(1,3).value=3

workbook.save(path)

